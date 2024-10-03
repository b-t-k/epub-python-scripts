#!python3
#coding: utf-8

import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from epub_functions import getinputfile, get_file, open_and_get
from epub_functions import clean_and_close 

###### START
# NOTE: Use inputFile for epubFile and Excel file

### get epub file
epubFile = None
# epubFile = "/Volumes/test files/Working/test.epub"
epubFile = getinputfile(epubFile)

### get excel file
inputFile = None
# inputFile="/Volumes/test files/Working/test.xlsx"
importFile = getinputfile(inputFile,".xlsx")

newFileEnding= "_alt"

################### NOTE ######################
# Revised version that searches every xhtml file 
# for image rather than relying on spreadsheet 
# to have the right file name. Should note missing images.
################### #### ######################

################### FUNCTION ######################
###################################################
### Function to search for image and replace alt text in files
def imagereplace(soup,imgsrc,newalttext):

    findimage=soup.find('img', src=imgsrc)
    if findimage==None:
        return("no_image") #not a string
        # print("MISSING IMAGE " + imgsrc)
        # return(soup) #not a string

    else: #replace alttext
        if newalttext != "PRESENTATION":
            findimage['alt']=newalttext
            # print(newalttext)
        else: # if marked as PRESENTATION then make alt blank and append role
            findimage['alt']=""
            findimage['role']="presentation"
    return(soup) #not a string

###################################################
###################################################


###### START ######
Exportfolder,epubfiles = open_and_get(epubFile)
print("export ",Exportfolder)

### get folder where image files are stored
### get the path of the folder
# imagefolder,imagefolderpath = get_file(Exportfolder,"imagefolder")
imagefolder_raw, image_folder_path = get_file(Exportfolder, "imagefolder")
imagefolder=imagefolder_raw + "/"

print("file count: " + str(len(epubfiles)))
filecount=len(epubfiles)

### Change to working directory
# print(os.getcwd())
os.chdir(Exportfolder)

### Open file and loop through to make list
wb = load_workbook(filename=importFile)
wb.sheetnames
print(wb.sheetnames)
sheet=wb.active

alttext = []

### Loop through and skip first line 
for desc in sheet.iter_rows(values_only=True, min_row=2):
# for desc in sheet.iter_rows(min_row=2):
    ### add desc to alttext list
    alttext.append(desc)
    # print(alttext)
    # print(alttext[3][2])
print("image count: " + str(len(alttext)))

### WRITE TO .XHTML FILES ###
### Loop though list of images
# print(alttext[0][1:3])
for image in alttext:
    execution_count=0
    imgname=image[1]
    newalttext=image[2] 
    # if newalttext is None:
    #     newalttext ="MISSING_ALT_TEXT"
    #     print(imgname + " — Missing alt text!")

    # # print(image[1])
    ### set image path
    if image[1] != None:
        imgsrc=imagefolder+imgname
        print(imgsrc)
    else:
        print("Missing Image Name")
        continue
    # print(imgsrc)
    # print(soup)
    ### For each image loop though all files and search
    for filename in epubfiles:        
        with open(filename, 'r') as inp:
            data = inp.read()
            soup=BeautifulSoup(data, 'xml')
            # print(filename)
            newdata = imagereplace(soup,imgsrc,newalttext)
        # exit()

        ### if img not found don't bother to write to file
        if newdata == "no_image":
            newdata = soup
            # print(newdata)
            execution_count += 1
            # print(execution_count)
            if execution_count >= filecount:
                print ("MISSING IMAGE: " + imgsrc)
            # continue

    ### open file as a binary to accept soup without changing it to a string
    ### changed to open  normally and then changin soup to string. Otherwise it screws with the formatting
        # with open(filename, 'wb') as output:
        with open(filename, 'w') as output:
            # output.write(newdata.prettify("UTF-8"))
            output.write(str(newdata))
            output.close()
    
    # exit()

clean_and_close(Exportfolder,newFileEnding)